import collections
import os
import sys

import openpyxl

import database
from truckmate_email import TruckmateEmail

REPORT_EMAILS = [
    'jwaltzjr@krclogistics.com'
]

class Rate(object):

    def __init__(self, tariff, customers, origin, destination, break_value, is_min, rate):
        self.tariff = tariff
        self.customers = customers
        self.origin = origin
        self.destination = destination
        self.break_value = break_value
        self.is_min = (is_min.strip() == 'True')
        self.rate = rate

    def __repr__(self):
        return 'Rate(tariff=%s, origin=%s, dest=%s, break=%s, rate=%s)' % (
            self.tariff,
            self.origin,
            self.destination,
            self.rate_break,
            self.rate
        )

    @property
    def three_digit_zip(self):
        if self.destination.isdigit():
            if 600 <= int(self.destination[:3]) <= 606:
                return 'CHICOMM'
            else:
                return self.destination[:3]
        elif self.destination == 'CHICOMM':
            return 'CHICOMM'
        elif self.destination in ['497LP', '497UP']:
            return '497'
        else:
            return 'OTHER'

    @property
    def rate_break(self):
        if self.is_min:
            return 'MIN'
        else:
            rounded_break = round(self.break_value / 100.0) * 100.0
            return rounded_break

class RateReport(object):

    def __init__(self, file_name, datab):
        sql_file_path = os.path.join(sys.path[0], file_name)
        self.sql_query = self.load_query_from_file(sql_file_path)
        self.dataset = self.fetch_data_from_db(self.sql_query, datab)
        self.split_data = self.split_dataset(self.dataset)

    def load_query_from_file(self, file_path):
        with open(file_path, 'r') as sql_file:
            return sql_file.read()

    def fetch_data_from_db(self, query, db):
        with db as datab:
            with datab.connection.cursor() as cursor:
                cursor.execute(query)
                return cursor.fetchall()

    def split_dataset(self, dataset):
        split_data = collections.defaultdict(
            lambda: {
                'breaks': set(),
                'rates': collections.defaultdict(list)
            }
        )

        for rate in dataset:
            for origin in self.get_origins(rate):
                rate_obj = Rate(rate.TARIFF, rate.CUSTOMERS, origin, rate.DESTINATION, rate.BREAK, rate.IS_MIN, rate.RATE)

                if rate_obj.rate_break not in split_data[rate_obj.three_digit_zip]['breaks']:
                    if not rate_obj.is_min:
                        split_data[rate_obj.three_digit_zip]['breaks'].add(rate_obj.rate_break)

                rate_tup = (rate_obj.tariff, rate_obj.customers, rate_obj.origin, rate_obj.destination)
                split_data[rate_obj.three_digit_zip]['rates'][rate_tup].append(rate_obj)

        return split_data

    def get_origins(self, rate):
        origins = []

        if rate.ORIGIN_MS:
            for origin in rate.ORIGIN_MS.split(', '):
                origins.append(origin)

        if rate.ORIGIN:
            origins.append(rate.ORIGIN)

        return origins

    def export_as_xlsx(self):
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.active)

        for zone in sorted(self.split_data.keys()):
            ws = wb.create_sheet(zone)
            self._excel_insert_titles(ws, zone)
            self._excel_insert_data(ws, zone)

        virtual_wb = openpyxl.writer.excel.save_virtual_workbook(wb)
        return virtual_wb

    def _excel_insert_titles(self, worksheet, zone):
        titles = {
            'A1': 'TARIFF',
            'B1': 'CUSTOMER',
            'C1': 'ORIGIN',
            'D1': 'DESTINATION',
            'E1': 'MIN'
        }

        row = 'F'
        for b in sorted(self.split_data[zone]['breaks']):
            cellname = row + str(1)
            titles[cellname] = b
            row = chr(ord(row) + 1)

        for cell, title in titles.items():
            worksheet[cell] = title

    def _excel_insert_data(self, worksheet, zone):
        current_row = 2
        for tariff_tup, rates in sorted(self.split_data[zone]['rates'].iteritems()):
            tariff, customers, origin, destination = tariff_tup
            worksheet.cell(row=current_row, column=1).value = tariff
            worksheet.cell(row=current_row, column=2).value = customers
            worksheet.cell(row=current_row, column=3).value = origin
            worksheet.cell(row=current_row, column=4).value = destination
            for rate in rates:
                current_column = self.find_column(worksheet, rate.rate_break)
                current_cell = worksheet.cell(row=current_row, column=current_column)
                current_cell.value = rate.rate
                current_cell.number_format = '#,##0.00'

            current_row += 1

    def find_column(self, worksheet, header):
        for cell in worksheet[1]:
            if cell.value == header:
                return cell.col_idx
        else:
            raise ValueError('No header found for %s' % header)

def main():
    rate_report = RateReport('ratereport.sql', database.truckmate)
    email_message = TruckmateEmail(
        REPORT_EMAILS,
        subject='Rate Report',
        attachments=[
            ('rate_report.xlsx', rate_report.export_as_xlsx())
        ]
    )
    email_message.send()

if __name__ == '__main__':
    main()
