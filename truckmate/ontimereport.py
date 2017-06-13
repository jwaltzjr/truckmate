import os
import StringIO
import sys

import openpyxl
import pandas

import database
from truckmate_email import TruckmateEmail

REPORT_EMAILS = [
    'jwaltzjr@krclogistics.com',
    'jwaltz@krclogistics.com',
    'dhendriksen@krclogistics.com',
    'djdevries@krclogistics.com',
    'tkatsahnias@krclogistics.com',
    'ekuhowski@krclogistics.com',
    'dpeach@krclogistics.com'
]

class OnTimeReport(object):

    class CalcColumns(object):

        @staticmethod
        def ontime_appt(delivery_date, rad):
            if rad:
                return delivery_date <= rad
            else:
                return True # Always on time if no due date

        @staticmethod
        def ontime_appt_realistic(rad, rpd, created_date, deliver_by):
            if rad and rpd and (rad > rpd) & (rpd >= created_date):
                return deliver_by <= rad
            elif not rad:
                return True
            else:
                return None

        @staticmethod
        def ontime_delv(arrived, deliver_by_end):
            if arrived:
                return arrived <= deliver_by_end
            else:
                return None

    def __init__(self, file_name, datab):
        self.sql_query = self.load_query_from_file(os.path.join(sys.path[0], file_name))
        self.dataset = self.fetch_data_from_db(self.sql_query, datab)
        self.apply_calculated_columns()

    @property
    def data_as_csv(self):
        virtual_csv = StringIO.StringIO()
        self.dataset.to_csv(virtual_csv)
        virtual_csv.seek(0)
        return virtual_csv.getvalue()

    def load_query_from_file(self, file_path):
        with open(file_path, 'r') as sql_file:
            return sql_file.read()

    def fetch_data_from_db(self, sql_query, datab):
        with datab as db:
            dataset = pandas.read_sql(
                sql_query,
                db.connection
            )
        return dataset

    def apply_calculated_columns(self):
        calculated_columns = {
            'ONTIME_APPT': (
                lambda row: self.__class__.CalcColumns.ontime_appt(
                    row['DELIVER_BY'].date(),
                    row['RAD']
                )
            ),
            'ONTIME_APPT_REALISTIC': (
                lambda row: self.__class__.CalcColumns.ontime_appt_realistic(
                    row['RAD'],
                    row['RPD'],
                    row['CREATED_TIME'].date(),
                    row['DELIVER_BY'].date()
                )
            ),
            'ONTIME_DELV': (
                lambda row: self.__class__.CalcColumns.ontime_delv(
                    row['ARRIVED'],
                    row['DELIVER_BY_END']
                )
            )
        }
        for name, calculation in calculated_columns.items():
            self.dataset[name] = self.dataset.apply(calculation, axis = 1)

    def get_dataset_of_averages(self):
        numeric_dataset = self.dataset[
            [
                'DELIVERY_WEEK',
                'DELIVERY_TERMINAL',
                'ONTIME_APPT',
                'ONTIME_APPT_REALISTIC',
                'ONTIME_DELV'
            ]
        ].apply(pandas.to_numeric, errors='ignore')

        averages = numeric_dataset.groupby(
            ['DELIVERY_WEEK', 'DELIVERY_TERMINAL']
        )[['ONTIME_APPT', 'ONTIME_APPT_REALISTIC', 'ONTIME_DELV']].mean()

        return averages

    def export_as_xlsx(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        data = self.get_dataset_of_averages()

        self.excel_insert_titles(ws)

        current_column = 1
        current_date = None
        for row in data.iloc[::-1].itertuples():
            if row.Index[0] != current_date:
                current_column += 1
            current_date = row.Index[0]
            self.excel_insert_data(ws, row, current_column, 'ONTIME_APPT')
            self.excel_insert_data(ws, row, current_column, 'ONTIME_APPT_REALISTIC', row_offset=13)
            self.excel_insert_data(ws, row, current_column, 'ONTIME_DELV', row_offset=26)

        self.excel_apply_styling(ws)

        virtual_wb = openpyxl.writer.excel.save_virtual_workbook(wb)
        return virtual_wb

    def excel_insert_titles(self, worksheet):
        titles = {
            'A1': 'DELIVERY WEEK',
            'A3': 'Appt On Time to RAD',
            'A4': 'STAL-TERM',
            'A5': 'COMM-TERM',
            'A6': 'KELL-TERM',
            'A7': 'LRFD2-TERM',
            'A8': 'FKSP-TERM',
            'A9': 'UPPN-TERM',
            'A10': 'UPPN2-TERM',
            'A11': 'HESS-TERM',
            'A12': 'HCIB-TERM',
            'A13': 'BUBM-TERM',
            'A14': 'RINF-TERM',
            'A16': 'Appt On Time to RAD (Realistic)',
            'A17': 'STAL-TERM',
            'A18': 'COMM-TERM',
            'A19': 'KELL-TERM',
            'A20': 'LRFD2-TERM',
            'A21': 'FKSP-TERM',
            'A22': 'UPPN-TERM',
            'A23': 'UPPN2-TERM',
            'A24': 'HESS-TERM',
            'A25': 'HCIB-TERM',
            'A26': 'BUBM-TERM',
            'A27': 'RINF-TERM',
            'A29': 'Delv On Time to Appt',
            'A30': 'STAL-TERM',
            'A31': 'COMM-TERM',
            'A32': 'KELL-TERM',
            'A33': 'LRFD2-TERM',
            'A34': 'FKSP-TERM',
            'A35': 'UPPN-TERM',
            'A36': 'UPPN2-TERM',
            'A37': 'HESS-TERM',
            'A38': 'HCIB-TERM',
            'A39': 'BUBM-TERM',
            'A40': 'RINF-TERM'
        }

        for cell, title in titles.items():
            worksheet[cell] = title

    def excel_insert_data(self, worksheet, ontime_week, column, ontime_field, row_offset=0):
        report_column = {
            'Delivery Week': worksheet.cell(row=1, column=column),
            'STAL-TERM': worksheet.cell(row=4+row_offset, column=column),
            'COMM-TERM': worksheet.cell(row=5+row_offset, column=column),
            'KELL-TERM': worksheet.cell(row=6+row_offset, column=column),
            'LRFD2-TERM': worksheet.cell(row=7+row_offset, column=column),
            'FKSP-TERM': worksheet.cell(row=8+row_offset, column=column),
            'UPPN-TERM': worksheet.cell(row=9+row_offset, column=column),
            'UPPN2-TERM': worksheet.cell(row=10+row_offset, column=column),
            'HESS-TERM': worksheet.cell(row=11+row_offset, column=column),
            'HCIB-TERM': worksheet.cell(row=12+row_offset, column=column),
            'BUBM-TERM': worksheet.cell(row=13+row_offset, column=column),
            'RINF-TERM': worksheet.cell(row=14+row_offset, column=column),
        }

        for key, cell in report_column.iteritems():
            if key != 'Delivery Week':
                cell.number_format = '0.00%'

        current_terminal = ontime_week.Index[1].strip()

        report_column['Delivery Week'].value = ontime_week.Index[0]
        report_column[current_terminal].value = getattr(ontime_week, ontime_field)

    def excel_apply_styling(self, worksheet):
        for spreadsheet_section in ['A', 3, 16, 29]:
            for cell in worksheet[spreadsheet_section]:
                cell.font = cell.font.copy(bold=True)
        for spreadsheet_cell in ['A3', 'A16', 'A29']:
            worksheet[spreadsheet_cell].font = worksheet[spreadsheet_cell].font.copy(underline='single')

def main():
    ontime_report = OnTimeReport('ontimereport.sql', database.truckmate)
    email_message = TruckmateEmail(
        REPORT_EMAILS,
        subject='On Time Report',
        attachments=[
            ('on_time_report.xlsx', ontime_report.export_as_xlsx()),
            ('on_time_report_data.csv', ontime_report.data_as_csv)
        ]
    )
    email_message.send()

if __name__ == '__main__':
    main()
