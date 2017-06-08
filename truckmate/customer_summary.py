import collections
import os
import sys

import openpyxl

import database
from truckmate_email import TruckmateEmail

REPORT_EMAILS = [
    'jwaltzjr@krclogistics.com'
]

SummaryReport = collections.namedtuple('SummaryReport', ['name', 'grouping'])

class CustomerSummary(object):

    def __init__(self, customer, datab):
        self.reports = [
            SummaryReport('bymonth', 'MONTH'),
            SummaryReport('bycommodity', 'COMMODITY')
        ]
        self.dataset = {}

        for report in self.reports:
            sql_file_path = os.path.join(sys.path[0], 'customer_summary_{}.sql'.format(report.name))
            self.sql_query = self.load_query_from_file(sql_file_path)
            # Must include customer twice to fill the query slots
            self.dataset[report.name] = self.fetch_data_from_db(self.sql_query, datab, customer, customer)

    def load_query_from_file(self, file_path):
        with open(file_path, 'r') as sql_file:
            return sql_file.read()

    def fetch_data_from_db(self, query, db, *args):
        with db as datab:
            with datab.connection.cursor() as cursor:
                cursor.execute(query, *args)
                return cursor.fetchall()

    def export_as_xlsx(self):
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.active)

        for report in self.reports:
            ws = wb.create_sheet(report.name)
            self._excel_summary_insert_titles(ws, report.grouping)
            current_row = 2
            for customer_data in self.dataset[report.name]:
                self._excel_summary_insert_data(ws, customer_data, current_row, report.grouping)
                current_row += 1
            self._excel_summary_apply_styling(ws)

        virtual_wb = openpyxl.writer.excel.save_virtual_workbook(wb)
        return virtual_wb

    def _excel_summary_insert_titles(self, worksheet, grouping):
        titles = {
            'A1': grouping,
            'B1': 'SHIPMENTS',
            'C1': 'TOTAL LBS',
            'D1': 'TOTAL PLT',
            'E1': 'TOTAL REVENUE',
            'F1': 'AVG LBS/SHIPMENT',
            'G1': 'AVG PLT/SHIPMENT',
            'H1': 'AVG REVENUE/SHIPMENT'
        }

        for cell, title in titles.items():
            worksheet[cell] = title


    def _excel_summary_insert_data(self, worksheet, customer_data, current_row, grouping):
        summary_row = {
            grouping: worksheet.cell(row=current_row, column=1),
            'SHIPMENTS': worksheet.cell(row=current_row, column=2),
            'TOTAL LBS': worksheet.cell(row=current_row, column=3),
            'TOTAL PLT': worksheet.cell(row=current_row, column=4),
            'TOTAL REVENUE': worksheet.cell(row=current_row, column=5),
            'AVG LBS/SHIPMENT': worksheet.cell(row=current_row, column=6),
            'AVG PLT/SHIPMENT': worksheet.cell(row=current_row, column=7),
            'AVG REVENUE/SHIPMENT': worksheet.cell(row=current_row, column=8)
        }

        for name in summary_row.keys():
            summary_row[name].value = getattr(customer_data, name)

    def _excel_summary_apply_styling(self, worksheet):
        bolded_sections = ['A', 1]
        # Add 'TOTAL' row to bolded items
        for cell in worksheet['A']:
            if cell.value == 'TOTAL':
                bolded_sections.append(cell.row)

        for spreadsheet_section in bolded_sections:
            for cell in worksheet[spreadsheet_section]:
                cell.font = cell.font.copy(bold=True)

def main():
    customer_summary = CustomerSummary(
        'NESTILDEKA',
        database.truckmate
    )
    email_message = TruckmateEmail(
        REPORT_EMAILS,
        subject='Customer Summary',
        attachments=[
            ('customer_summary.xlsx', customer_summary.export_as_xlsx())
        ]
    )
    email_message.send()

if __name__ == '__main__':
    main()
