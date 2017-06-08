import os
import sys

import openpyxl

import database
from truckmate_email import TruckmateEmail

REPORT_EMAILS = [
    'jwaltzjr@krclogistics.com',
    'jwaltz@krclogistics.com',
    'dhendriksen@krclogistics.com',
    'djdevries@krclogistics.com',
    'tkatsahnias@krclogistics.com',
    'ekuhowski@krclogistics.com',
    'dpeach@krclogistics.com'
]

class TonnageReport(object):

    def __init__(self, file_name, datab):
        sql_file_path = os.path.join(sys.path[0], file_name)
        self.sql_query = self.load_query_from_file(sql_file_path)
        self.dataset = self.fetch_data_from_db(self.sql_query, datab)

    def load_query_from_file(self, file_path):
        with open(file_path, 'r') as sql_file:
            return sql_file.read()

    def fetch_data_from_db(self, query, db):
        with db as datab:
            with datab.connection.cursor() as cursor:
                cursor.execute(query)
                return cursor.fetchall()

    def export_as_xlsx(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        self._excel_insert_titles(ws)

        current_column = 2
        for tonnage_week in self.dataset:
            self._excel_insert_data(ws, tonnage_week, current_column)
            current_column += 1

        self._excel_apply_styling(ws)

        virtual_wb = openpyxl.writer.excel.save_virtual_workbook(wb)
        return virtual_wb

    def _excel_insert_titles(self, worksheet):
        titles = {
            'A1': 'DELIVERY WEEK',

            'A3': '# ORDERS 10',
            'A4': '# ORDERS 11',
            'A5': '# ORDERS 12',
            'A6': '# ORDERS 13',
            'A7': '# ORDERS 14',
            'A8': '# ORDERS 15',
            'A9': '# ORDERS UNDEF',
            'A10': '# ORDERS TOTAL',

            'A12': 'WEIGHT 10',
            'A13': 'WEIGHT 11',
            'A14': 'WEIGHT 12',
            'A15': 'WEIGHT 13',
            'A16': 'WEIGHT 14',
            'A17': 'WEIGHT 15',
            'A18': 'WEIGHT UNDEF',
            'A19': 'WEIGHT TOTAL',

            'A21': 'AVG WEIGHT 10',
            'A22': 'AVG WEIGHT 11',
            'A23': 'AVG WEIGHT 12',
            'A24': 'AVG WEIGHT 13',
            'A25': 'AVG WEIGHT 14',
            'A26': 'AVG WEIGHT 15',
            'A27': 'AVG WEIGHT UNDEF',
            'A28': 'AVG WEIGHT TOTAL',

            'A30': 'PALLETS 10',
            'A31': 'PALLETS 11',
            'A32': 'PALLETS 12',
            'A33': 'PALLETS 13',
            'A34': 'PALLETS 14',
            'A35': 'PALLETS 15',
            'A36': 'PALLETS UNDEF',
            'A37': 'PALLETS TOTAL',

            'A39': 'AVG PALLETS 10',
            'A40': 'AVG PALLETS 11',
            'A41': 'AVG PALLETS 12',
            'A42': 'AVG PALLETS 13',
            'A43': 'AVG PALLETS 14',
            'A44': 'AVG PALLETS 15',
            'A45': 'AVG PALLETS UNDEF',
            'A46': 'AVG PALLETS TOTAL',

            'A48': 'POSITIONS 10',
            'A49': 'POSITIONS 11',
            'A50': 'POSITIONS 12',
            'A51': 'POSITIONS 13',
            'A52': 'POSITIONS 14',
            'A53': 'POSITIONS 15',
            'A54': 'POSITIONS UNDEF',
            'A55': 'POSITIONS TOTAL',

            'A57': 'AVG POSITIONS 10',
            'A58': 'AVG POSITIONS 11',
            'A59': 'AVG POSITIONS 12',
            'A60': 'AVG POSITIONS 13',
            'A61': 'AVG POSITIONS 14',
            'A62': 'AVG POSITIONS 15',
            'A63': 'AVG POSITIONS UNDEF',
            'A64': 'AVG POSITIONS TOTAL'
        }

        for cell, title in titles.items():
            worksheet[cell] = title


    def _excel_insert_data(self, worksheet, tonnage_week, column):
        tonnage_week_column = {
            'DELIVERY_WEEK': worksheet.cell(row=1, column=column),

            'NUM_ORDERS_10': worksheet.cell(row=3, column=column),
            'NUM_ORDERS_11': worksheet.cell(row=4, column=column),
            'NUM_ORDERS_12': worksheet.cell(row=5, column=column),
            'NUM_ORDERS_13': worksheet.cell(row=6, column=column),
            'NUM_ORDERS_14': worksheet.cell(row=7, column=column),
            'NUM_ORDERS_15': worksheet.cell(row=8, column=column),
            'NUM_ORDERS_UNDEF': worksheet.cell(row=9, column=column),
            'NUM_ORDERS': worksheet.cell(row=10, column=column),

            'WEIGHT_10': worksheet.cell(row=12, column=column),
            'WEIGHT_11': worksheet.cell(row=13, column=column),
            'WEIGHT_12': worksheet.cell(row=14, column=column),
            'WEIGHT_13': worksheet.cell(row=15, column=column),
            'WEIGHT_14': worksheet.cell(row=16, column=column),
            'WEIGHT_15': worksheet.cell(row=17, column=column),
            'WEIGHT_UNDEF': worksheet.cell(row=18, column=column),
            'WEIGHT': worksheet.cell(row=19, column=column),

            'AVG_WEIGHT_10': worksheet.cell(row=21, column=column),
            'AVG_WEIGHT_11': worksheet.cell(row=22, column=column),
            'AVG_WEIGHT_12': worksheet.cell(row=23, column=column),
            'AVG_WEIGHT_13': worksheet.cell(row=24, column=column),
            'AVG_WEIGHT_14': worksheet.cell(row=25, column=column),
            'AVG_WEIGHT_15': worksheet.cell(row=26, column=column),
            'AVG_WEIGHT_UNDEF': worksheet.cell(row=27, column=column),
            'AVG_WEIGHT': worksheet.cell(row=28, column=column),

            'PALLETS_10': worksheet.cell(row=30, column=column),
            'PALLETS_11': worksheet.cell(row=31, column=column),
            'PALLETS_12': worksheet.cell(row=32, column=column),
            'PALLETS_13': worksheet.cell(row=33, column=column),
            'PALLETS_14': worksheet.cell(row=34, column=column),
            'PALLETS_15': worksheet.cell(row=35, column=column),
            'PALLETS_UNDEF': worksheet.cell(row=36, column=column),
            'PALLETS': worksheet.cell(row=37, column=column),

            'AVG_PALLETS_10': worksheet.cell(row=39, column=column),
            'AVG_PALLETS_11': worksheet.cell(row=40, column=column),
            'AVG_PALLETS_12': worksheet.cell(row=41, column=column),
            'AVG_PALLETS_13': worksheet.cell(row=42, column=column),
            'AVG_PALLETS_14': worksheet.cell(row=43, column=column),
            'AVG_PALLETS_15': worksheet.cell(row=44, column=column),
            'AVG_PALLETS_UNDEF': worksheet.cell(row=45, column=column),
            'AVG_PALLETS': worksheet.cell(row=46, column=column),

            'POSITIONS_10': worksheet.cell(row=48, column=column),
            'POSITIONS_11': worksheet.cell(row=49, column=column),
            'POSITIONS_12': worksheet.cell(row=50, column=column),
            'POSITIONS_13': worksheet.cell(row=51, column=column),
            'POSITIONS_14': worksheet.cell(row=52, column=column),
            'POSITIONS_15': worksheet.cell(row=53, column=column),
            'POSITIONS_UNDEF': worksheet.cell(row=54, column=column),
            'POSITIONS': worksheet.cell(row=55, column=column),

            'AVG_POSITIONS_10': worksheet.cell(row=57, column=column),
            'AVG_POSITIONS_11': worksheet.cell(row=58, column=column),
            'AVG_POSITIONS_12': worksheet.cell(row=59, column=column),
            'AVG_POSITIONS_13': worksheet.cell(row=60, column=column),
            'AVG_POSITIONS_14': worksheet.cell(row=61, column=column),
            'AVG_POSITIONS_15': worksheet.cell(row=62, column=column),
            'AVG_POSITIONS_UNDEF': worksheet.cell(row=63, column=column),
            'AVG_POSITIONS': worksheet.cell(row=64, column=column)
        }

        for key, cell in tonnage_week_column.iteritems():
            if key != 'title':
                cell.number_format = '#,##0'

        for name in tonnage_week_column.keys():
            tonnage_week_column[name].value = getattr(tonnage_week, name)

    def _excel_apply_styling(self, worksheet):
        for spreadsheet_section in ['A', 1, 10, 19, 28, 37, 46, 55, 64]:
            for cell in worksheet[spreadsheet_section]:
                cell.font = cell.font.copy(bold=True)
        for cell in ['A9', 'A18', 'A27', 'A36', 'A45', 'A54', 'A63']:
            worksheet[cell].font = worksheet[cell].font.copy(underline='single')

def main():
    tonnage_report = TonnageReport('tonnage.sql', database.truckmate)
    email_message = TruckmateEmail(
        REPORT_EMAILS,
        subject='Weekly Tonnage',
        attachments=[
            ('weekly_tonnage.xlsx', tonnage_report.export_as_xlsx())
        ]
    )
    email_message.send()

if __name__ == '__main__':
    main()
